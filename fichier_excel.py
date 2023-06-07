import openpyxl


oct1 = openpyxl.load_workbook("octobre.xlsx", data_only=True)
oct2 = openpyxl.load_workbook("novembre.xlsx", data_only=True)
oct3 = openpyxl.load_workbook("decembre.xlsx", data_only=True)


def ajouter_data_depuis_oct(oct, d):
    sheet = oct.active
    for row in range(2, sheet.max_row):
        nom_article = sheet.cell(row, 1).value
        if not nom_article:
            break
        total_ventes = sheet.cell(row, 4).value
        if d.get(nom_article):
            d[nom_article].append(total_ventes)
        else:
            d[nom_article] = [total_ventes]



donnees = {}
ajouter_data_depuis_oct(oct1, donnees)
ajouter_data_depuis_oct(oct2, donnees)
ajouter_data_depuis_oct(oct3, donnees)
print(donnees)

# ------- Creer un fichier excel -------
 
oct_sortie = openpyxl.Workbook()
sheet = oct_sortie.active
sheet["A1"] = "Article"
sheet["B1"] = "Octobre"
sheet["C1"] = "Novembre"
sheet["D1"] = "Decembre"

row = 2
for i in donnees.items():
    nom_article = i[0]
    ventes =i[1]
    sheet.cell(row, 1).value = nom_article
    for j in range(0, len(ventes)):
        sheet.cell(row, 2+j).value = ventes[j]
    row += 1

# ------ Création d'un graphique ------

chart_ref = openpyxl.chart.Reference(sheet, min_col=2, min_row=2, max_col=sheet.max_column, max_row=2)
chart_serie = openpyxl.chart.Series(chart_ref, title="Total ventes £")
chart = openpyxl.chart.BarChart()
chart.title ="Evolution du prix des pommes"
chart.append(chart_serie)

sheet.add_chart(chart, "J2")

chart_ref = openpyxl.chart.Reference(sheet, min_col=2, min_row=6, max_col=sheet.max_column, max_row=6)
chart_serie = openpyxl.chart.Series(chart_ref, title="Total ventes £")
chart = openpyxl.chart.BarChart()
chart.title ="Evolution du prix des ananas"
chart.append(chart_serie)

sheet.add_chart(chart, "A10")

chart_ref = openpyxl.chart.Reference(sheet, min_col=2, min_row=4, max_col=sheet.max_column, max_row=4)
chart_serie = openpyxl.chart.Series(chart_ref, title="Total ventes £")
chart = openpyxl.chart.BarChart()
chart.title ="Evolution du prix des bananes"
chart.append(chart_serie)

sheet.add_chart(chart, "J18")

chart_ref = openpyxl.chart.Reference(sheet, min_col=2, min_row=5, max_col=sheet.max_column, max_row=5)
chart_serie = openpyxl.chart.Series(chart_ref, title="Total ventes £")
chart = openpyxl.chart.BarChart()
chart.title ="Evolution du prix des mangues"
chart.append(chart_serie)

sheet.add_chart(chart, "A26")

oct_sortie.save("Total_ventes_trimestre.xlsx")
